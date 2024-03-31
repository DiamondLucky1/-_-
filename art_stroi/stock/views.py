import openpyxl
import xlwt
from django.contrib.auth.views import redirect_to_login
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib import messages
from django.views import View
from django.contrib.auth.decorators import login_required


from .forms import *
from .models import *

from unicodedata import category
from rest_framework.views import APIView
from django.contrib.auth import authenticate, login, logout

#Главная страница, Вход
class IndexView(APIView):
    def get(self, request, *args, **kwargs):
        form = LoginUserForm()
        product = Product.objects.all()
        return render(request, 'stock/index.html', {'title': 'Вход', 'form': form, 'product': product})

    def post(self, request, *args, **kwargs):
        username = request.POST.get('username') #Получаем введенное имя пользователя из запроса
        password = request.POST.get('password') #Получаем введенный пароль пользователя из запроса
        user = authenticate(request, username=username, password=password)# Аутентифицируем пользователя
        if user is not None: # Если пользователь аутентифицирован успешно
            login(request, user) # Осуществляем вход пользователя
            return redirect('main') # Перенаправляем на главную страницу
        else: # Если пользователя не удалось аутентифицировать
            messages.error(request, 'Неверное имя пользователя или пароль') # Добавляем сообщение об ошибке в request
            return redirect('index') # Перенаправляем на страницу входа


#Выход пользователя с системы
class Logout(View):
    def get(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            logout(request)
        return redirect('index')
    

#Вывод ошибки , если пользователь не вошел в систе
def permission_denied(request, exception):
    return render(request, '403.html', status=403)


#Главная стр.после входа в систему
def main(request):
     if not request.user.is_anonymous:
         products = Product.objects.all()
         categories = Category.objects.all()
         manufacturers = Manufacturer.objects.all()
         branches = Branch.objects.all()
         return render(request, 'stock/main.html', {'title':'Главная страница',
        'products':products, 'categories': categories, 'manufacturers': manufacturers, 'branches':
        branches})
     else:
        return redirect('/')


#Выборка продукта, применение фильтров
def product_search(request):
    if request.user.is_authenticated:
        products = Product.objects.all() # Получаем все продукты из базы

        # Применяем фильтр по наименованию, если был введен поисковый запрос
        search_query = request.GET.get('search_query')
        if search_query:
            products = products.filter(title__icontains=search_query)

        # Применяем фильтры по категории, производителю
        category_filter = request.GET.get('category')
        if category_filter:
            products = products.filter(category=category_filter)

        # Применяем фильтры по категории, филиалу
        manufacturer_filter = request.GET.get('manufacturer')
        if manufacturer_filter:
            products = products.filter(manufacturer=manufacturer_filter)

        # Также реализуем фильтр по филиалу и другие фильтры

        return render(request, 'stock/main.html', {'products': products})
    return redirect('index')

 # Закупка товара

def purchase(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            form = ApplicationForm(request.POST)
            if form.is_valid():
                name_of_manager = form.cleaned_data['name_of_manager']
                title_of_product = form.cleaned_data['title_of_product']
                amount_of_product = form.cleaned_data['amount_of_product']
                characteristic_of_product = form.cleaned_data['characteristic_of_product']
                new_aplication = application(name_of_manager=name_of_manager, title_of_product=title_of_product,
                                             amount_of_product=amount_of_product, characteristic_of_product=characteristic_of_product)
                new_aplication.save()
                messages.success(request, 'Товар успешно добавлен на закуп!')
                return redirect(reverse('form'))

        else:
            form = ApplicationForm()
            return render(request, 'stock/form.html', {'form': form})
    else:
        return redirect('index')

## на примере класса
# class Formpost(APIView):
#     def get(self, request, *args, **kwargs):
#         return render(request, 'stock/form.html', {'title': 'Форма'})
#
#     def post(self, request, *args, **kwargs):
#         name_of_manager = request.POST.get('imya_menegera')
#         title_of_product = request.POST.get('naz_tovar')
#         amount_of_product = request.POST.get('kolvo_tovar')
#         characteristic_of_product = request.POST.get('comment')
#
#         if not name_of_manager or not title_of_product or not amount_of_product:
#             return HttpResponseBadRequest("Please fill in all the fields")
#
#         new_model = application()
#         new_model.name_of_manager = name_of_manager
#         new_model.title_of_product = title_of_product
#         new_model.amount_of_product = amount_of_product
#         new_model.characteristic_of_product = characteristic_of_product
#
#         new_model.save()
#         return redirect(reverse('form'))
#         # return render(request, 'stock/form.html', {'title': 'Форма'})


##Скачивание из БД оставшего товара в Excel ф-л
#
# def downloads_product(request):
#     products = Product.objects.all()
#     print(products)
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.append(['Наименование товара', 'Цена', 'Характеристика', 'Количество', 'Категория'])
#     for product in products:
#         ws.append(
#             [product.title, product.price, product.characteristic, product.amount, product.category.category_name])
#
#     response = HttpResponse(content_type='application/ms-excel')
#     response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
#     wb.save(response)
#     return response


# def make_report1(request):
#     products = Product.objects.all()
#     print(products)
#     wb = openpyxl.Workbook() ##создаем новую рабочую книгу Excel
#     ws = wb.active ##получаем активный лист из созданной рабочей книги
#     ws.append(['Наименование товара', 'Цена', 'Характеристика', 'Количество', 'Категория'] ) ##добавляем заголовки столбцов в лист Excel
#
#     ##Добавляем данные о товарах
#     for product in products:
#         ws.append([product.title,
#                    product.price, product.characteristic, product.amount,
#                    product.category.category_name])##создается список и добавляется в лист Excel все перечисленные поля
#     response = HttpResponse(content_type='application/ms-excel')#создает новый объект HttpResponse с \
#                                                                     # \типом контента 'application/ms-excel', обозначающим, что это Excel файл
#
#     response['Content-Disposition'] = 'attachment; filename="products.xlsx"'# устанавливает заголовок\
#                                                                                 # Content-Disposition для HTTP ответа\
#                                                         #attachment сообщает браузеру, что ответ должен быть сохранен как файл
#
#     wb.save(response) ##сохраняет созданную рабочую книгу Excel `wb` в объекте `response`
#
#     return response



##Скачивание из БД оставшего товара в Excel ф-л как вариант с классом APIView
class MakeReport1(APIView):
    def get(self, request):
        products = Product.objects.all()
        wb = xlwt.Workbook(encoding="utf-8") ##создаем новую рабочую книгу Excel
        ws = wb.add_sheet("Список имеющихся товаров")# создается новый лист с указанным названием
        headers = ["Наименование товара", "Характеристика", "Цена", "Количество", "Производитель", "Категория", "Филиал"]
        for col, header in enumerate(headers):
            ws.write(0, col, header)
        row_num = 1
        for entry in products:
            ws.write(row_num, 0, entry.title)# запись значения `header` в ячейку в строке с индексом 0 (первая строка) и
                                            # столбце с индексом `col` (полученном в результате `enumerate(headers)`).
            ws.write(row_num, 1, entry.characteristic)
            ws.write(row_num, 2, entry.price)
            ws.write(row_num, 3, entry.amount)
            ws.write(row_num, 4, entry.manufacturer.manufacturer_name)
            ws.write(row_num, 5, entry.category.category_name)
            ws.write(row_num, 6, entry.branch.branch_name)
            row_num += 1
        file_path = "report.xls" #устанавливается имя файла, в который будет сохранен Excel-файл.
        wb.save(file_path) #Этот метод сохраняет созданный документ Excel по указанному пути.
        with open(file_path, 'rb') as f: #открывается файл `report.xls` в режиме чтения байтов (`'rb'`).
                                            # Файл будет прочитан как байтовый поток.
            response = HttpResponse(f.read(), content_type="application/vnd.ms-excel")#Используется для создания объекта
                                            # `HttpResponse` с содержимым, считанным из файла.
            response['Content-Disposition'] = 'attachment; filename="report.xls"'#Устанавливается заголовок `Content-Disposition`,
                                            # указывающий, что файл должен быть предложен для скачивания с именем файла "report.xls".
        return response  #Завершает обработку запроса и возвращает сформированный объект `response`,
                            # который содержит файл Excel в качестве прикрепленного документа.



#Скачать список товара на закуп в excel
def make_report2(request):
    appl = application.objects.all()
    print(appl)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Имя менеджера', 'Название продукта', 'Количество', 'Характеристика'])
    for entry in appl:
        ws.append([entry.name_of_manager, entry.title_of_product, entry.amount_of_product, entry.characteristic_of_product])
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename ="report.xlsx"'
    wb.save(response)
    return response

